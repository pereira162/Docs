# Excel Extractor for RAG System
# Extrator completo para arquivos .xlsx com máxima extração de informações

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
from typing import Dict, List, Any, Optional
import json
import re
from datetime import datetime
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("excel_extractor")

class ExcelExtractor:
    """
    Extrator completo para arquivos Excel (.xlsx) que maximiza a extração de informações
    incluindo planilhas, nomes, conteúdo, metadados, fórmulas e formatação.
    """
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls']
        
    def extract_from_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Extrai o máximo de informações de um arquivo Excel
        
        Returns:
            Dict contendo todas as informações extraídas organizadas para RAG
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
            
        if file_path.suffix.lower() not in self.supported_formats:
            raise ValueError(f"Formato não suportado: {file_path.suffix}")
        
        logger.info(f"🔍 Extraindo dados do Excel: {file_path.name}")
        
        results = {
            "file_info": self._extract_file_info(file_path),
            "workbook_info": {},
            "worksheets": {},
            "summary": {},
            "content_for_rag": {},
            "metadata": {}
        }
        
        try:
            # Usar openpyxl para metadados e estrutura avançada
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            results["workbook_info"] = self._extract_workbook_info(workbook)
            
            # Usar pandas para leitura eficiente de dados
            all_sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            
            # Extrair informações detalhadas de cada planilha
            for sheet_name, df in all_sheets.items():
                logger.info(f"📊 Processando planilha: {sheet_name}")
                
                worksheet_info = self._extract_worksheet_info(
                    sheet_name, df, workbook[sheet_name] if sheet_name in workbook.sheetnames else None
                )
                results["worksheets"][sheet_name] = worksheet_info
            
            # Criar resumo geral
            results["summary"] = self._create_summary(results)
            
            # Preparar conteúdo otimizado para RAG
            results["content_for_rag"] = self._prepare_rag_content(results)
            
            # Metadados finais
            results["metadata"] = self._extract_metadata(results)
            
            logger.info(f"✅ Extração concluída: {len(results['worksheets'])} planilhas processadas")
            
        except Exception as e:
            logger.error(f"❌ Erro na extração: {e}")
            results["error"] = str(e)
            
        return results
    
    def _extract_file_info(self, file_path: Path) -> Dict[str, Any]:
        """Extrai informações básicas do arquivo"""
        stat = file_path.stat()
        return {
            "filename": file_path.name,
            "size_bytes": stat.st_size,
            "size_mb": round(stat.st_size / (1024 * 1024), 2),
            "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
            "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "extension": file_path.suffix.lower()
        }
    
    def _extract_workbook_info(self, workbook) -> Dict[str, Any]:
        """Extrai informações do workbook usando openpyxl"""
        info = {
            "total_sheets": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "active_sheet": workbook.active.title if workbook.active else None,
            "properties": {}
        }
        
        # Propriedades do documento
        if hasattr(workbook, 'properties'):
            props = workbook.properties
            info["properties"] = {
                "title": getattr(props, 'title', None),
                "creator": getattr(props, 'creator', None),
                "description": getattr(props, 'description', None),
                "subject": getattr(props, 'subject', None),
                "keywords": getattr(props, 'keywords', None),
                "category": getattr(props, 'category', None),
                "created": getattr(props, 'created', None),
                "modified": getattr(props, 'modified', None)
            }
        
        return info
    
    def _extract_worksheet_info(self, sheet_name: str, df: pd.DataFrame, worksheet=None) -> Dict[str, Any]:
        """Extrai informações detalhadas de uma planilha"""
        info = {
            "name": sheet_name,
            "shape": {"rows": len(df), "columns": len(df.columns)},
            "columns": list(df.columns),
            "data_types": df.dtypes.to_dict(),
            "has_header": True,  # Assumimos que há cabeçalho
            "empty_cells": df.isnull().sum().sum(),
            "total_cells": len(df) * len(df.columns),
            "data_preview": {},
            "content_analysis": {},
            "raw_data": [],
            "formatted_data": [],
            "tables": [],
            "charts": [],
            "formulas": []
        }
        
        # Análise de conteúdo
        info["content_analysis"] = self._analyze_content(df)
        
        # Preview dos dados (primeiras e últimas linhas)
        info["data_preview"] = {
            "head": df.head(3).to_dict(orient='records') if len(df) > 0 else [],
            "tail": df.tail(3).to_dict(orient='records') if len(df) > 3 else [],
            "sample": df.sample(min(5, len(df))).to_dict(orient='records') if len(df) > 0 else []
        }
        
        # Dados completos (limitado para evitar sobrecarga)
        if len(df) <= 1000:  # Limite para evitar arquivos muito grandes
            info["raw_data"] = df.to_dict(orient='records')
        else:
            info["raw_data"] = df.head(1000).to_dict(orient='records')
            info["note"] = f"Apenas primeiras 1000 linhas mostradas de {len(df)} total"
        
        # Dados formatados para texto
        info["formatted_data"] = self._format_data_for_text(df, sheet_name)
        
        # Detectar tabelas estruturadas
        info["tables"] = self._detect_tables(df)
        
        # Extrair fórmulas e informações avançadas se worksheet disponível
        if worksheet:
            info["formulas"] = self._extract_formulas(worksheet)
            info["charts"] = self._extract_charts(worksheet)
            info["comments"] = self._extract_comments(worksheet)
            info["merged_cells"] = self._extract_merged_cells(worksheet)
        
        return info
    
    def _analyze_content(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Analisa o conteúdo da planilha para extrair insights"""
        analysis = {
            "numeric_columns": [],
            "text_columns": [],
            "date_columns": [],
            "boolean_columns": [],
            "empty_columns": [],
            "unique_values": {},
            "statistics": {},
            "patterns": []
        }
        
        for col in df.columns:
            col_data = df[col].dropna()
            
            if len(col_data) == 0:
                analysis["empty_columns"].append(col)
                continue
                
            # Classificar tipo de coluna
            if pd.api.types.is_numeric_dtype(col_data):
                analysis["numeric_columns"].append(col)
                analysis["statistics"][col] = {
                    "mean": float(col_data.mean()) if not col_data.empty else None,
                    "min": float(col_data.min()) if not col_data.empty else None,
                    "max": float(col_data.max()) if not col_data.empty else None,
                    "std": float(col_data.std()) if not col_data.empty else None
                }
            elif pd.api.types.is_datetime64_any_dtype(col_data):
                analysis["date_columns"].append(col)
            elif pd.api.types.is_bool_dtype(col_data):
                analysis["boolean_columns"].append(col)
            else:
                analysis["text_columns"].append(col)
            
            # Valores únicos (limitado)
            unique_vals = col_data.unique()
            if len(unique_vals) <= 50:  # Limite para evitar sobrecarga
                analysis["unique_values"][col] = list(unique_vals)
            else:
                analysis["unique_values"][col] = f"{len(unique_vals)} valores únicos"
        
        # Detectar padrões
        analysis["patterns"] = self._detect_patterns(df)
        
        return analysis
    
    def _detect_patterns(self, df: pd.DataFrame) -> List[str]:
        """Detecta padrões nos dados"""
        patterns = []
        
        try:
            # Converter nomes de colunas para string e depois para lower
            column_names = [str(col) for col in df.columns]
            
            # Detectar se é tabela de dados transacionais
            if any(str(col).lower() in ['data', 'date', 'timestamp'] for col in column_names):
                patterns.append("Dados transacionais com timestamps")
            
            # Detectar se é relatório financeiro
            financial_keywords = ['valor', 'preco', 'custo', 'receita', 'despesa', 'total', 'subtotal']
            if any(any(keyword in str(col).lower() for keyword in financial_keywords) for col in column_names):
                patterns.append("Dados financeiros")
            
            # Detectar se é inventário/estoque
            inventory_keywords = ['produto', 'item', 'codigo', 'quantidade', 'estoque']
            if any(any(keyword in str(col).lower() for keyword in inventory_keywords) for col in column_names):
                patterns.append("Dados de inventário/estoque")
            
            # Detectar se é dados de pessoas
            people_keywords = ['nome', 'email', 'telefone', 'cpf', 'cnpj', 'endereco']
            if any(any(keyword in str(col).lower() for keyword in people_keywords) for col in column_names):
                patterns.append("Dados pessoais/clientes")
            
            # Detectar dados educacionais
            education_keywords = ['aluno', 'nota', 'disciplina', 'turma', 'atividade', 'prova', 'trabalho', 'escola']
            if any(any(keyword in str(col).lower() for keyword in education_keywords) for col in column_names):
                patterns.append("Dados educacionais")
                
        except Exception as e:
            self.logger.warning(f"Erro na detecção de padrões: {e}")
            patterns.append("Dados gerais")
        
        return patterns if patterns else ["Dados gerais"]
    
    def _format_data_for_text(self, df: pd.DataFrame, sheet_name: str) -> List[str]:
        """Formata dados para texto legível em RAG"""
        formatted = []
        
        # Adicionar cabeçalho da planilha
        formatted.append(f"## Planilha: {sheet_name}")
        formatted.append(f"Dimensões: {len(df)} linhas x {len(df.columns)} colunas")
        formatted.append("")
        
        # Listar colunas
        formatted.append("### Colunas:")
        for i, col in enumerate(df.columns, 1):
            formatted.append(f"{i}. {col}")
        formatted.append("")
        
        # Amostrar dados (primeiras linhas)
        if len(df) > 0:
            formatted.append("### Dados de exemplo:")
            for idx, row in df.head(10).iterrows():
                row_text = " | ".join([f"{col}: {str(row[col])}" for col in df.columns if pd.notna(row[col])])
                formatted.append(f"Linha {idx + 1}: {row_text}")
            formatted.append("")
        
        return formatted
    
    def _detect_tables(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Detecta tabelas estruturadas dentro da planilha"""
        tables = []
        
        # Tabela principal (dados completos)
        if len(df) > 0:
            tables.append({
                "type": "main_table",
                "description": "Tabela principal com todos os dados",
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": list(df.columns),
                "sample_data": df.head(3).to_dict(orient='records')
            })
        
        # Detectar sub-tabelas por grupos vazios
        # (implementação básica - pode ser expandida)
        
        return tables
    
    def _extract_formulas(self, worksheet) -> List[Dict[str, Any]]:
        """Extrai fórmulas da planilha"""
        formulas = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    formulas.append({
                        "cell": cell.coordinate,
                        "formula": cell.value,
                        "result": cell.displayed_value if hasattr(cell, 'displayed_value') else None
                    })
        
        return formulas
    
    def _extract_charts(self, worksheet) -> List[Dict[str, Any]]:
        """Extrai informações sobre gráficos"""
        charts = []
        
        for chart in worksheet._charts:
            charts.append({
                "type": type(chart).__name__,
                "title": getattr(chart, 'title', None),
                "position": str(chart.anchor) if hasattr(chart, 'anchor') else None
            })
        
        return charts
    
    def _extract_comments(self, worksheet) -> List[Dict[str, Any]]:
        """Extrai comentários das células"""
        comments = []
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.comment:
                    comments.append({
                        "cell": cell.coordinate,
                        "comment": cell.comment.text,
                        "author": cell.comment.author if hasattr(cell.comment, 'author') else None
                    })
        
        return comments
    
    def _extract_merged_cells(self, worksheet) -> List[str]:
        """Extrai informações sobre células mescladas"""
        return [str(merged_range) for merged_range in worksheet.merged_cells.ranges]
    
    def _create_summary(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """Cria resumo geral do arquivo Excel"""
        summary = {
            "total_worksheets": len(results["worksheets"]),
            "worksheet_names": list(results["worksheets"].keys()),
            "total_data_rows": 0,
            "total_columns": 0,
            "content_types": set(),
            "has_formulas": False,
            "has_charts": False,
            "has_comments": False,
            "key_insights": []
        }
        
        for sheet_name, sheet_info in results["worksheets"].items():
            summary["total_data_rows"] += sheet_info["shape"]["rows"]
            summary["total_columns"] += sheet_info["shape"]["columns"]
            
            # Adicionar tipos de conteúdo
            summary["content_types"].update(sheet_info["content_analysis"]["patterns"])
            
            # Verificar recursos avançados
            if sheet_info.get("formulas"):
                summary["has_formulas"] = True
            if sheet_info.get("charts"):
                summary["has_charts"] = True
            if sheet_info.get("comments"):
                summary["has_comments"] = True
        
        summary["content_types"] = list(summary["content_types"])
        
        # Insights
        if summary["total_worksheets"] > 1:
            summary["key_insights"].append(f"Arquivo com múltiplas planilhas ({summary['total_worksheets']} planilhas)")
        
        if summary["has_formulas"]:
            summary["key_insights"].append("Contém fórmulas e cálculos")
        
        if summary["has_charts"]:
            summary["key_insights"].append("Contém gráficos e visualizações")
        
        return summary
    
    def _prepare_rag_content(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """Prepara conteúdo otimizado para RAG"""
        rag_content = {
            "document_text": [],
            "structured_data": {},
            "searchable_chunks": [],
            "metadata_text": []
        }
        
        # Texto principal do documento
        file_info = results["file_info"]
        rag_content["document_text"].append(f"# Arquivo Excel: {file_info['filename']}")
        rag_content["document_text"].append(f"Tamanho: {file_info['size_mb']} MB")
        rag_content["document_text"].append("")
        
        # Resumo
        summary = results["summary"]
        rag_content["document_text"].append(f"## Resumo")
        rag_content["document_text"].append(f"- Total de planilhas: {summary['total_worksheets']}")
        rag_content["document_text"].append(f"- Planilhas: {', '.join(summary['worksheet_names'])}")
        rag_content["document_text"].append(f"- Total de linhas de dados: {summary['total_data_rows']}")
        rag_content["document_text"].append("")
        
        # Conteúdo de cada planilha
        for sheet_name, sheet_info in results["worksheets"].items():
            rag_content["document_text"].extend(sheet_info["formatted_data"])
            rag_content["document_text"].append("")
            
            # Dados estruturados
            rag_content["structured_data"][sheet_name] = {
                "columns": sheet_info["columns"],
                "sample_data": sheet_info["data_preview"]["head"],
                "patterns": sheet_info["content_analysis"]["patterns"]
            }
            
            # Chunks pesquisáveis
            chunk_text = f"Planilha {sheet_name}: " + " ".join([
                f"Coluna {col}" for col in sheet_info["columns"]
            ])
            rag_content["searchable_chunks"].append({
                "text": chunk_text,
                "source": f"sheet_{sheet_name}",
                "type": "column_names"
            })
            
            # Adicionar dados como chunks
            for i, row in enumerate(sheet_info["data_preview"]["head"]):
                row_text = " ".join([f"{k}: {v}" for k, v in row.items() if pd.notna(v)])
                rag_content["searchable_chunks"].append({
                    "text": row_text,
                    "source": f"sheet_{sheet_name}_row_{i}",
                    "type": "data_row"
                })
        
        return rag_content
    
    def _extract_metadata(self, results: Dict[str, Any]) -> Dict[str, Any]:
        """Extrai metadados finais"""
        return {
            "extraction_timestamp": datetime.now().isoformat(),
            "extractor_version": "1.0.0",
            "file_type": "excel",
            "total_worksheets": len(results["worksheets"]),
            "processing_method": "pandas + openpyxl",
            "features_extracted": [
                "worksheet_names", "data_content", "formulas", 
                "charts", "comments", "merged_cells", "data_types"
            ]
        }

def main():
    """Função de teste"""
    print("🚀 EXCEL EXTRACTOR - TESTE")
    print("=" * 50)
    
    # Procurar arquivo Excel na pasta atual
    excel_files = list(Path(".").glob("*.xlsx")) + list(Path(".").glob("*.xls"))
    
    if not excel_files:
        print("❌ Nenhum arquivo Excel (.xlsx/.xls) encontrado na pasta atual")
        print("📋 Coloque um arquivo Excel na pasta e tente novamente")
        return False
    
    # Usar o primeiro arquivo encontrado
    excel_file = excel_files[0]
    print(f"📄 Arquivo encontrado: {excel_file.name}")
    
    try:
        extractor = ExcelExtractor()
        results = extractor.extract_from_excel(str(excel_file))
        
        # Salvar resultados
        output_dir = Path("excel_outputs")
        output_dir.mkdir(exist_ok=True)
        
        # JSON completo
        json_file = output_dir / f"{excel_file.stem}_extraction.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False, default=str)
        
        # Markdown para RAG
        md_file = output_dir / f"{excel_file.stem}_rag_content.md"
        with open(md_file, 'w', encoding='utf-8') as f:
            f.write("\n".join(results["content_for_rag"]["document_text"]))
        
        # Chunks para busca
        chunks_file = output_dir / f"{excel_file.stem}_chunks.json"
        with open(chunks_file, 'w', encoding='utf-8') as f:
            json.dump(results["content_for_rag"]["searchable_chunks"], f, indent=2, ensure_ascii=False)
        
        print("\n" + "="*50)
        print("✅ EXTRAÇÃO CONCLUÍDA COM SUCESSO!")
        print(f"📄 Planilhas processadas: {len(results['worksheets'])}")
        print(f"📊 Total de linhas: {results['summary']['total_data_rows']}")
        print(f"🗂️ Chunks gerados: {len(results['content_for_rag']['searchable_chunks'])}")
        print("\n💾 Arquivos salvos:")
        print(f"  📄 {json_file}")
        print(f"  📝 {md_file}")
        print(f"  🔍 {chunks_file}")
        
        # Mostrar resumo das planilhas
        print("\n📋 PLANILHAS ENCONTRADAS:")
        for name, info in results["worksheets"].items():
            print(f"  • {name}: {info['shape']['rows']} linhas, {info['shape']['columns']} colunas")
            if info["content_analysis"]["patterns"]:
                print(f"    Padrões: {', '.join(info['content_analysis']['patterns'])}")
        
        return True
        
    except Exception as e:
        print(f"❌ ERRO: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = main()
    if success:
        print("\n🎉 TESTE CONCLUÍDO COM SUCESSO!")
    else:
        print("\n❌ TESTE FALHOU")
